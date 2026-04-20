"""
Organisation Classifier (4-Category)
======================================
Classifies organisations into Research / Health / Government / Industry
using a three-tier approach for maximum accuracy and reproducibility:

  Tier 1 - ROR metadata (instant, deterministic):
    Maps ROR 'type' field directly:
      education    -> Research
      healthcare   -> Health
      government   -> Government
      company      -> Industry
      nonprofit    -> Research (research foundations, learned societies)
      facility     -> Government
      archive      -> Government

  Tier 2 - Wikidata (deterministic):
    Searches for the org by name, retrieves 'instance of' (P31) claims,
    and classifies based on Wikidata entity types.

  Tier 3 - Claude LLM (cached for reproducibility):
    For remaining unclassified orgs, queries the Anthropic API with
    temperature=0 for a deterministic classification. Results are cached
    so re-runs produce identical output.

Input:  Output of extract_unique_orgs.py (has Organisation_Name, ROR_ID, Type, etc.)
Output: Same data with added Classification and Classification_Source columns.

Setup:
  1. Same config.ini as other scripts, with added section:
       [anthropic]
       api_key = sk-ant-...
  2. pip install openpyxl requests

Usage:
  python classify_orgs.py <unique_orgs.xlsx> [output.xlsx] [config.ini]
"""

import sys
import re
import json
import time
import configparser
from pathlib import Path
import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

SCRIPT_DIR = Path(__file__).resolve().parent

WIKIDATA_SEARCH = "https://www.wikidata.org/w/api.php"
WIKIDATA_ENTITY = "https://www.wikidata.org/w/api.php"
ANTHROPIC_API = "https://api.anthropic.com/v1/messages"

# ── ROR type mapping (Tier 1) ──────────────────────────────────────────

ROR_TYPE_MAP = {
    "education":  "Research",
    "healthcare": "Health",
    "government": "Government",
    "company":    "Industry",
    "nonprofit":  "Research",
    "facility":   "Government",
    "archive":    "Government",
}

# ── Wikidata instance_of -> classification mapping (Tier 2) ─────────

WIKIDATA_RESEARCH = {
    "Q3918",       # university
    "Q875538",     # public university
    "Q902104",     # private university
    "Q15936437",   # research university
    "Q38723",      # higher education institution
    "Q31855",      # research institute
    "Q2467461",    # research center
    "Q4830453",    # business school (academic)
    "Q1371037",    # polytechnic
    "Q189004",     # college
    "Q2385804",    # educational institution
    "Q23002039",   # scientific institute
    "Q1664720",    # institute of technology
    "Q1188663",    # technical university
    "Q2659904",    # medical school
    "Q751108",     # academy of sciences
    "Q3354859",    # collegiate university
    "Q10498148",   # research centre
    "Q62078547",   # research organization
    "Q955824",     # learned society
}

WIKIDATA_HEALTH = {
    "Q16917",      # hospital
    "Q179661",     # public hospital
    "Q7315155",    # research hospital
    "Q1244922",    # teaching hospital
    "Q4260475",    # medical centre
    "Q1774898",    # community health centre
    "Q697175",     # health system
    "Q2145977",    # clinic
    "Q1391494",    # mental health centre
    "Q180958",     # psychiatric hospital
    "Q1127957",    # rehabilitation centre
    "Q837346",     # children's hospital
    "Q2889518",    # veterans hospital
    "Q64578911",   # health service
    "Q1774838",    # health care provider
    "Q55659167",   # hospital network
}

WIKIDATA_GOVERNMENT = {
    "Q327333",     # government agency
    "Q7188",       # government
    "Q35798",      # executive branch
    "Q1328899",    # ministry
    "Q637846",     # government office
    "Q1548775",    # public body
    "Q2824523",    # statutory authority
    "Q4508",       # municipality
    "Q15284",      # local government
    "Q1063239",    # national laboratory
    "Q431289",     # public research institution
    "Q895526",     # regulatory agency
    "Q1785733",    # government department
}

WIKIDATA_INDUSTRY = {
    "Q891723",     # public company
    "Q6881511",    # enterprise
    "Q783794",     # business firm
    "Q1589009",    # privately held company
    "Q161726",     # multinational corporation
    "Q134161",     # joint-stock company
    "Q5225895",    # corporation
    "Q786820",     # limited liability company
    "Q2912172",    # pharmaceutical company
    "Q18388277",   # technology company
    "Q4830453",    # business enterprise
}


# ── Config ──────────────────────────────────────────────────────────────

def load_config(config_path=None):
    if config_path is None:
        config_path = SCRIPT_DIR / "config.ini"
    else:
        config_path = Path(config_path).resolve()

    if not config_path.exists():
        print(f"ERROR: Config file not found at {config_path}")
        sys.exit(1)

    config = configparser.ConfigParser()
    config.read(config_path)
    email = config.get("crossref", "email", fallback=None)
    if not email or email.strip() == "your_email@example.com":
        print(f"ERROR: Please set your real email in {config_path}")
        sys.exit(1)

    api_key = config.get("anthropic", "api_key", fallback="")

    return {
        "email": email.strip(),
        "delay": config.getfloat("crossref", "delay", fallback=1),
        "save_every": config.getint("crossref", "save_every", fallback=50),
        "max_retries": config.getint("crossref", "max_retries", fallback=3),
        "api_key": api_key.strip(),
    }


# ── Cache ───────────────────────────────────────────────────────────────

def load_cache(cache_path):
    if cache_path.exists():
        with open(cache_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_cache(cache, cache_path):
    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)


# ── Tier 1: ROR metadata ───────────────────────────────────────────────

def classify_by_ror(ror_type):
    """Classify using ROR type field. Returns classification or None."""
    if not ror_type:
        return None
    ror_type_lower = ror_type.lower()
    return ROR_TYPE_MAP.get(ror_type_lower)


# ── Tier 2: Wikidata ───────────────────────────────────────────────────

def search_wikidata_entity(org_name, session, max_retries=3):
    """Search Wikidata for an org by name. Returns entity ID or None."""
    params = {
        "action": "wbsearchentities",
        "search": org_name,
        "language": "en",
        "limit": 3,
        "format": "json",
    }

    for attempt in range(1, max_retries + 1):
        try:
            resp = session.get(WIKIDATA_SEARCH, params=params, timeout=30)
            if resp.status_code == 429:
                time.sleep(2 ** attempt)
                continue
            resp.raise_for_status()
            data = resp.json()
            results = data.get("search", [])
            if results:
                return results[0].get("id")
            return None
        except requests.exceptions.RequestException:
            if attempt < max_retries:
                time.sleep(2 ** attempt)
            else:
                return None
    return None


def get_wikidata_instance_of(entity_id, session, max_retries=3):
    """Get P31 (instance of) claims for a Wikidata entity."""
    params = {
        "action": "wbgetclaims",
        "entity": entity_id,
        "property": "P31",
        "format": "json",
    }

    for attempt in range(1, max_retries + 1):
        try:
            resp = session.get(WIKIDATA_ENTITY, params=params, timeout=30)
            if resp.status_code == 429:
                time.sleep(2 ** attempt)
                continue
            resp.raise_for_status()
            data = resp.json()
            claims = data.get("claims", {}).get("P31", [])
            q_ids = []
            for claim in claims:
                mainsnak = claim.get("mainsnak", {})
                datavalue = mainsnak.get("datavalue", {})
                value = datavalue.get("value", {})
                qid = value.get("id")
                if qid:
                    q_ids.append(qid)
            return q_ids
        except requests.exceptions.RequestException:
            if attempt < max_retries:
                time.sleep(2 ** attempt)
            else:
                return []
    return []


def classify_by_wikidata(org_name, session, cache, max_retries=3):
    """
    Search Wikidata, get P31, classify.
    Priority: Research > Health > Government > Industry
    Returns (classification, entity_id, p31_types) or (None, None, []).
    """
    cache_key = org_name.lower().strip()
    if cache_key in cache:
        cached = cache[cache_key]
        return cached.get("classification"), cached.get("entity_id"), cached.get("p31_types", [])

    entity_id = search_wikidata_entity(org_name, session, max_retries)
    if not entity_id:
        cache[cache_key] = {"classification": None, "entity_id": None, "p31_types": []}
        return None, None, []

    time.sleep(0.5)

    p31_ids = get_wikidata_instance_of(entity_id, session, max_retries)
    if not p31_ids:
        cache[cache_key] = {"classification": None, "entity_id": entity_id, "p31_types": []}
        return None, entity_id, []

    p31_set = set(p31_ids)

    # Priority order: Research > Health > Government > Industry
    if p31_set & WIKIDATA_RESEARCH:
        result = "Research"
    elif p31_set & WIKIDATA_HEALTH:
        result = "Health"
    elif p31_set & WIKIDATA_GOVERNMENT:
        result = "Government"
    elif p31_set & WIKIDATA_INDUSTRY:
        result = "Industry"
    else:
        result = None

    cache[cache_key] = {
        "classification": result,
        "entity_id": entity_id,
        "p31_types": p31_ids,
    }
    return result, entity_id, p31_ids


# ── Tier 3: Claude LLM ─────────────────────────────────────────────────

LLM_SYSTEM_PROMPT = """You are classifying research organisations into exactly one of four categories.

Categories:
- RESEARCH: Universities, research institutes, academic centres, learned societies, academic publishers, research foundations, medical research institutes
- HEALTH: Hospitals, health services, clinical networks, medical centres, mental health services, rehabilitation centres, aged care facilities, health departments of hospitals
- GOVERNMENT: Government departments, agencies, public health authorities, national laboratories, statutory bodies, regulatory bodies, councils, government-funded research agencies (e.g., CSIRO, NIH, CNRS)
- INDUSTRY: Private companies, corporations, commercial entities, industry R&D labs, consulting firms, pharmaceutical companies, technology companies

Rules:
- Teaching hospitals and university-affiliated hospitals -> HEALTH (not Research)
- Standalone medical research institutes (e.g., Walter and Eliza Hall Institute) -> RESEARCH
- Public hospitals -> HEALTH
- Government health departments (policy, not clinical) -> GOVERNMENT
- Pharmaceutical companies -> INDUSTRY
- Hospital-based research centres -> HEALTH
- Non-profit research foundations -> RESEARCH
- Government-funded national research agencies (e.g., CSIRO, NIH) -> GOVERNMENT

Respond with ONLY the category name (RESEARCH, HEALTH, GOVERNMENT, or INDUSTRY) followed by a pipe character and a one-sentence justification. Example:
HEALTH | This is a public teaching hospital providing clinical services.

Do not include any other text."""


def classify_by_llm(org_name, country, ror_type, api_key, session, cache, max_retries=3):
    """Classify using Claude API with temperature=0 for reproducibility."""
    cache_key = org_name.lower().strip()
    if cache_key in cache:
        cached = cache[cache_key]
        return cached.get("classification"), cached.get("justification", "")

    if not api_key:
        return None, "[No API key configured]"

    user_msg = f"Organisation: {org_name}"
    if country:
        user_msg += f"\nCountry: {country}"
    if ror_type:
        user_msg += f"\nType from registry: {ror_type}"

    payload = {
        "model": "claude-sonnet-4-20250514",
        "max_tokens": 150,
        "temperature": 0,
        "system": LLM_SYSTEM_PROMPT,
        "messages": [{"role": "user", "content": user_msg}],
    }

    headers = {
        "Content-Type": "application/json",
        "x-api-key": api_key,
        "anthropic-version": "2023-06-01",
    }

    for attempt in range(1, max_retries + 1):
        try:
            resp = session.post(ANTHROPIC_API, json=payload, headers=headers, timeout=60)
            if resp.status_code == 429:
                wait = min(2 ** attempt * 5, 60)
                print(f"    LLM rate limited. Waiting {wait}s...")
                time.sleep(wait)
                continue
            resp.raise_for_status()
            data = resp.json()

            text = ""
            for block in data.get("content", []):
                if block.get("type") == "text":
                    text += block.get("text", "")

            text = text.strip()

            if "|" in text:
                category, justification = text.split("|", 1)
                category = category.strip().upper()
                justification = justification.strip()
            else:
                category = text.strip().upper()
                justification = ""

            category_map = {
                "RESEARCH": "Research",
                "HEALTH": "Health",
                "GOVERNMENT": "Government",
                "INDUSTRY": "Industry",
            }
            classification = category_map.get(category)

            if classification:
                cache[cache_key] = {
                    "classification": classification,
                    "justification": justification,
                }
                return classification, justification

            cache[cache_key] = {
                "classification": None,
                "justification": f"[Unexpected LLM response: {text}]",
            }
            return None, f"[Unexpected LLM response: {text}]"

        except requests.exceptions.RequestException as e:
            if attempt < max_retries:
                time.sleep(2 ** attempt)
            else:
                return None, f"[LLM error: {e}]"

    return None, "[LLM failed after retries]"


# ── Main ────────────────────────────────────────────────────────────────

def main(input_file, output_file=None, config_path=None):
    cfg = load_config(config_path)
    has_api_key = bool(cfg["api_key"])
    print(f"Config: email={cfg['email']}, LLM tier: {'enabled' if has_api_key else 'disabled (no API key)'}")

    # Resolve paths
    input_path = Path(input_file)
    if not input_path.is_absolute():
        if not input_path.exists():
            fallback = SCRIPT_DIR / input_path
            if fallback.exists():
                input_path = fallback
    input_path = input_path.resolve()

    if not input_path.exists():
        print(f"ERROR: Input file not found: {input_path}")
        sys.exit(1)

    if output_file is None:
        output_path = input_path.parent / f"{input_path.stem}_classified.xlsx"
    else:
        output_path = Path(output_file).resolve()

    wikidata_cache_path = output_path.parent / f"{input_path.stem}_wikidata_cache.json"
    llm_cache_path = output_path.parent / f"{input_path.stem}_llm_cache.json"

    print(f"Input:  {input_path}")
    print(f"Output: {output_path}")

    wikidata_cache = load_cache(wikidata_cache_path)
    llm_cache = load_cache(llm_cache_path)
    if wikidata_cache:
        print(f"  Loaded Wikidata cache: {len(wikidata_cache)} entries")
    if llm_cache:
        print(f"  Loaded LLM cache: {len(llm_cache)} entries")

    # --- Read input ---
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    headers = {cell.value: cell.column for cell in ws[1]}
    name_col = headers.get("Organisation_Name")
    ror_id_col = headers.get("ROR_ID")
    country_col = headers.get("Country")
    city_col = headers.get("City")
    type_col = headers.get("Type")
    variants_col = headers.get("Raw_Variants")
    author_count_col = headers.get("Author_Count")
    authors_col = headers.get("Authors")
    doi_count_col = headers.get("DOI_Count")
    dois_col = headers.get("DOIs")

    if name_col is None:
        print("ERROR: Could not find 'Organisation_Name' column.")
        sys.exit(1)

    orgs = []
    for row in range(2, ws.max_row + 1):
        org = {
            "name": str(ws.cell(row=row, column=name_col).value or "").strip(),
            "ror_id": str(ws.cell(row=row, column=ror_id_col).value or "").strip() if ror_id_col else "",
            "country": str(ws.cell(row=row, column=country_col).value or "").strip() if country_col else "",
            "city": str(ws.cell(row=row, column=city_col).value or "").strip() if city_col else "",
            "type": str(ws.cell(row=row, column=type_col).value or "").strip() if type_col else "",
            "variants": str(ws.cell(row=row, column=variants_col).value or "").strip() if variants_col else "",
            "author_count": ws.cell(row=row, column=author_count_col).value if author_count_col else 0,
            "authors": str(ws.cell(row=row, column=authors_col).value or "").strip() if authors_col else "",
            "doi_count": ws.cell(row=row, column=doi_count_col).value if doi_count_col else 0,
            "dois": str(ws.cell(row=row, column=dois_col).value or "").strip() if dois_col else "",
        }
        if org["name"]:
            orgs.append(org)

    print(f"\nFound {len(orgs)} organisations to classify.\n")

    # --- Tier 1: ROR metadata ---
    print("--- Tier 1: ROR metadata ---")
    tier1_count = 0
    tier2_needed = []

    for org in orgs:
        result = classify_by_ror(org["type"])
        if result:
            org["classification"] = result
            org["source"] = "ROR"
            org["justification"] = f"ROR type: {org['type']}"
            tier1_count += 1
        else:
            tier2_needed.append(org)

    print(f"  Classified: {tier1_count}")
    print(f"  Remaining:  {len(tier2_needed)}")

    # --- Tier 2: Wikidata ---
    print(f"\n--- Tier 2: Wikidata ---")
    session = requests.Session()
    session.headers.update({"User-Agent": f"OrgClassifier/1.0 (mailto:{cfg['email']})"})

    tier2_count = 0
    tier3_needed = []

    try:
        for i, org in enumerate(tier2_needed):
            display = org["name"][:60].encode("ascii", errors="replace").decode("ascii")
            print(f"  [{i+1}/{len(tier2_needed)}] {display}")

            result, entity_id, p31_types = classify_by_wikidata(
                org["name"], session, wikidata_cache, cfg["max_retries"]
            )

            if result:
                org["classification"] = result
                org["source"] = "Wikidata"
                org["justification"] = f"Wikidata {entity_id}, P31: {', '.join(p31_types[:3])}"
                tier2_count += 1
            else:
                tier3_needed.append(org)

            if (i + 1) % cfg["save_every"] == 0:
                save_cache(wikidata_cache, wikidata_cache_path)

            time.sleep(cfg["delay"])

    except KeyboardInterrupt:
        print(f"\n>> Interrupted at Wikidata tier. Saving caches...")
        save_cache(wikidata_cache, wikidata_cache_path)

    save_cache(wikidata_cache, wikidata_cache_path)

    print(f"  Classified: {tier2_count}")
    print(f"  Remaining:  {len(tier3_needed)}")

    # --- Tier 3: Claude LLM ---
    print(f"\n--- Tier 3: Claude LLM ---")
    tier3_count = 0
    unclassified = []

    if not has_api_key:
        print("  Skipped (no API key in config.ini [anthropic] section).")
        for org in tier3_needed:
            org["classification"] = "Unclassified"
            org["source"] = "None"
            org["justification"] = "No API key for LLM tier"
        unclassified = tier3_needed
    else:
        try:
            for i, org in enumerate(tier3_needed):
                display = org["name"][:60].encode("ascii", errors="replace").decode("ascii")
                print(f"  [{i+1}/{len(tier3_needed)}] {display}")

                result, justification = classify_by_llm(
                    org["name"], org["country"], org["type"],
                    cfg["api_key"], session, llm_cache, cfg["max_retries"]
                )

                if result:
                    org["classification"] = result
                    org["source"] = "LLM"
                    org["justification"] = justification
                    tier3_count += 1
                else:
                    org["classification"] = "Unclassified"
                    org["source"] = "None"
                    org["justification"] = justification
                    unclassified.append(org)

                if (i + 1) % cfg["save_every"] == 0:
                    save_cache(llm_cache, llm_cache_path)

                time.sleep(cfg["delay"])

        except KeyboardInterrupt:
            print(f"\n>> Interrupted at LLM tier. Saving caches...")

        save_cache(llm_cache, llm_cache_path)

    print(f"  Classified: {tier3_count}")
    print(f"  Unclassified: {len(unclassified)}")

    # --- Summary ---
    from collections import Counter
    class_counts = Counter(org.get("classification", "Unclassified") for org in orgs)
    source_counts = Counter(org.get("source", "None") for org in orgs)

    print(f"\n--- Summary ---")
    print(f"  By classification:")
    for cls in ["Research", "Health", "Government", "Industry", "Unclassified"]:
        print(f"    {cls}: {class_counts.get(cls, 0)}")
    print(f"  By source:")
    for src in ["ROR", "Wikidata", "LLM", "None"]:
        print(f"    {src}: {source_counts.get(src, 0)}")

    # --- Write output ---
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Classified Organisations"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_align = Alignment(horizontal="left", vertical="center")

    col_headers = [
        "Organisation_Name", "Classification", "Classification_Source",
        "Justification", "ROR_ID", "Country", "City", "ROR_Type",
        "Raw_Variants", "Author_Count", "Authors", "DOI_Count", "DOIs"
    ]
    for col_idx, hdr in enumerate(col_headers, 1):
        cell = out_ws.cell(row=1, column=col_idx, value=hdr)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    class_fills = {
        "Research":     PatternFill("solid", fgColor="D6EAF8"),  # light blue
        "Health":       PatternFill("solid", fgColor="E8DAEF"),  # light purple
        "Government":   PatternFill("solid", fgColor="D5F5E3"),  # light green
        "Industry":     PatternFill("solid", fgColor="FADBD8"),  # light red
        "Unclassified": PatternFill("solid", fgColor="F9E79F"),  # light yellow
    }

    data_font = Font(name="Arial")
    wrap_align = Alignment(wrap_text=True, vertical="top")

    for row_idx, org in enumerate(orgs, 2):
        out_ws.cell(row=row_idx, column=1, value=org["name"]).font = data_font

        cls = org.get("classification", "Unclassified")
        cell = out_ws.cell(row=row_idx, column=2, value=cls)
        cell.font = Font(name="Arial", bold=True)
        cell.fill = class_fills.get(cls, PatternFill())

        out_ws.cell(row=row_idx, column=3, value=org.get("source", "")).font = data_font
        cell = out_ws.cell(row=row_idx, column=4, value=org.get("justification", ""))
        cell.font = data_font
        cell.alignment = wrap_align
        out_ws.cell(row=row_idx, column=5, value=org["ror_id"]).font = data_font
        out_ws.cell(row=row_idx, column=6, value=org["country"]).font = data_font
        out_ws.cell(row=row_idx, column=7, value=org["city"]).font = data_font
        out_ws.cell(row=row_idx, column=8, value=org["type"]).font = data_font

        cell = out_ws.cell(row=row_idx, column=9, value=org["variants"])
        cell.font = data_font
        cell.alignment = wrap_align

        out_ws.cell(row=row_idx, column=10, value=org["author_count"]).font = data_font
        cell = out_ws.cell(row=row_idx, column=11, value=org["authors"])
        cell.font = data_font
        cell.alignment = wrap_align
        out_ws.cell(row=row_idx, column=12, value=org["doi_count"]).font = data_font
        cell = out_ws.cell(row=row_idx, column=13, value=org["dois"])
        cell.font = data_font
        cell.alignment = wrap_align

    out_ws.column_dimensions['A'].width = 40
    out_ws.column_dimensions['B'].width = 16
    out_ws.column_dimensions['C'].width = 20
    out_ws.column_dimensions['D'].width = 50
    out_ws.column_dimensions['E'].width = 35
    out_ws.column_dimensions['F'].width = 18
    out_ws.column_dimensions['G'].width = 18
    out_ws.column_dimensions['H'].width = 15
    out_ws.column_dimensions['I'].width = 50
    out_ws.column_dimensions['J'].width = 14
    out_ws.column_dimensions['K'].width = 50
    out_ws.column_dimensions['L'].width = 12
    out_ws.column_dimensions['M'].width = 60
    out_ws.freeze_panes = "A2"
    out_ws.auto_filter.ref = f"A1:M{out_ws.max_row}"

    out_wb.save(output_path)
    print(f"\nDone. Output saved to: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python classify_orgs.py <unique_orgs.xlsx> [output.xlsx] [config.ini]")
        sys.exit(1)
    in_arg = sys.argv[1]
    out_arg = sys.argv[2] if len(sys.argv) > 2 else None
    cfg_arg = sys.argv[3] if len(sys.argv) > 3 else None
    main(in_arg, out_arg, cfg_arg)